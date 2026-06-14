"""
Simplified API endpoints for RAG-based document analysis
No preview step - direct upload to analysis
"""

from flask import Blueprint, request, jsonify, send_file
import os
import io
import uuid
import shutil
from pathlib import Path
from werkzeug.utils import secure_filename
from app.rag_analyzer import make_analyzer, docx_to_html
from app.document_processor import DocumentProcessor

api_bp = Blueprint('api', __name__, url_prefix='/api')

# Store session metadata
_sessions = {}

# Allowed formats for the document under review & reference sources (knowledge base)
ALLOWED_EXTENSIONS = {'docx', 'doc', 'pdf', 'html'}
# Allowed formats for rule files — markdown/text/word permitted
ALLOWED_RULE_EXTENSIONS = {'md', 'txt', 'docx', 'doc'}

# ---------------------------------------------------------------------------
# Preset library (saved reference / rule sets for reuse)
# ---------------------------------------------------------------------------
_RAG_DATA_DIR = Path(__file__).parent.parent / 'rag' / 'data'
REFERENCE_PRESET_DIR = _RAG_DATA_DIR / 'reference_documents'
RULES_PRESET_DIR = _RAG_DATA_DIR / 'rules'
# Default presets bundled with the system cannot be deleted
_PROTECTED_PRESETS = {'quy_dinh_chung.md', '86nd.docx'}


def _preset_dir(kind: str) -> Path:
    """Return the preset directory for the given kind ('reference' | 'rule')."""
    return REFERENCE_PRESET_DIR if kind == 'reference' else RULES_PRESET_DIR


def _list_presets(kind: str, allowed_ext: set) -> list:
    """List the valid preset files in the library."""
    d = _preset_dir(kind)
    if not d.exists():
        return []
    items = []
    for p in sorted(d.iterdir()):
        if p.is_file() and _ext(p.name) in allowed_ext:
            items.append({
                'name': p.name,
                'size': p.stat().st_size,
                'protected': p.name in _PROTECTED_PRESETS,
            })
    return items


def _save_paths_as_presets(saved_paths: list, target_dir: Path) -> None:
    """
    Copy the uploaded files (paths inside upload_dir) into the preset library.
    Strip the 'ref_N_' / 'rule_N_' prefix to restore the original name.
    """
    import re
    target_dir.mkdir(parents=True, exist_ok=True)
    for path in saved_paths:
        base = os.path.basename(path)
        clean = re.sub(r'^(?:ref|rule)_\d+_', '', base)
        dest = target_dir / clean
        try:
            if not dest.exists():
                shutil.copy2(path, dest)
                print(f"[API] Lưu preset: {clean}")
        except Exception as e:  # noqa: BLE001
            print(f"[API] Không lưu được preset '{clean}': {e}")


def _resolve_preset_paths(kind: str, names: list, allowed_ext: set) -> list:
    """
    Convert a list of preset names → absolute paths (safe, guards against path traversal).
    Skip invalid / non-existent names.
    """
    d = _preset_dir(kind)
    paths = []
    for raw in names:
        name = os.path.basename((raw or '').strip())  # block '../'
        if not name or _ext(name) not in allowed_ext:
            continue
        candidate = d / name
        if candidate.exists() and candidate.is_file():
            paths.append(str(candidate))
    return paths


def _ext(filename):
    return filename.rsplit('.', 1)[1].lower() if '.' in filename else ''


def allowed_file(filename):
    return _ext(filename) in ALLOWED_EXTENSIONS


def allowed_rule_file(filename):
    return _ext(filename) in ALLOWED_RULE_EXTENSIONS


@api_bp.route('/upload', methods=['POST'])
def upload_and_analyze():
    """
    Upload documents and run analysis immediately
    
    Files expected:
    - mainDocument: DOCX file to analyze (required)
    - referenceDocuments: DOCX/HTML files for building the RAG knowledge base
      (optional — falls back to bundled NĐ 86 if none)
    - ruleDocuments: MD/TXT/DOCX files with the checking rules
      (optional — falls back to bundled quy_dinh_chung.md if none)
    """
    try:
        # Validate main document
        if 'mainDocument' not in request.files:
            return jsonify({'error': 'Main document (mainDocument) is required'}), 400
        
        main_doc = request.files['mainDocument']
        if not main_doc.filename or not allowed_file(main_doc.filename):
            return jsonify({'error': 'Main document must be .docx format'}), 400
        
        # Get reference and rule documents
        #   - referenceDocuments: reference sources → build the RAG index (knowledge base)
        #   - ruleDocuments: rules → inject into the audit prompt (NOT indexed)
        reference_docs = request.files.getlist('referenceDocuments')
        rule_docs = request.files.getlist('ruleDocuments')
        # historyDocuments: previously approved YCKT → lookup store for the Q&A chatbot
        history_docs = request.files.getlist('historyDocuments')

        # Create session directory
        session_id = str(uuid.uuid4())
        upload_dir = os.path.join(
            os.path.dirname(__file__), '..', 'uploads', session_id
        )
        os.makedirs(upload_dir, exist_ok=True)
        
        try:
            # Save main document
            main_filename = secure_filename(main_doc.filename or 'main.docx')
            main_path = os.path.join(upload_dir, main_filename)
            main_doc.save(main_path)
            print(f"[API] Saved main document: {main_filename}")
            
            # Save reference documents (reference sources → RAG index)
            ref_paths = []
            for ref_doc in reference_docs:
                if ref_doc and ref_doc.filename and allowed_file(ref_doc.filename):
                    ref_filename = secure_filename(ref_doc.filename)
                    ref_path = os.path.join(upload_dir, f'ref_{len(ref_paths)}_{ref_filename}')
                    ref_doc.save(ref_path)
                    ref_paths.append(ref_path)
                    print(f"[API] Saved reference: {ref_filename}")

            # Save rule documents (rules → audit prompt)
            rule_paths = []
            for rule_doc in rule_docs:
                if rule_doc and rule_doc.filename and allowed_rule_file(rule_doc.filename):
                    rule_filename = secure_filename(rule_doc.filename)
                    rule_path = os.path.join(upload_dir, f'rule_{len(rule_paths)}_{rule_filename}')
                    rule_doc.save(rule_path)
                    rule_paths.append(rule_path)
                    print(f"[API] Saved rule: {rule_filename}")

            # Save history documents (old YCKT → lookup store for the chatbot)
            # Keep the ORIGINAL NAME (history_names) for display/citation — the saved
            # filename has lost its Vietnamese diacritics via secure_filename and gained
            # a hist_N_ prefix.
            history_paths = []
            history_names = {}
            for hist_doc in history_docs:
                if hist_doc and hist_doc.filename and allowed_file(hist_doc.filename):
                    hist_filename = secure_filename(hist_doc.filename)
                    hist_path = os.path.join(
                        upload_dir, f'hist_{len(history_paths)}_{hist_filename}'
                    )
                    hist_doc.save(hist_path)
                    history_paths.append(hist_path)
                    history_names[hist_path] = hist_doc.filename  # user's original name
                    print(f"[API] Saved history YCKT: {hist_doc.filename}")

            # --- Save the JUST-UPLOADED files as presets (before merging existing presets) ---
            if request.form.get('savePresets', '').lower() == 'true':
                _save_paths_as_presets(ref_paths, REFERENCE_PRESET_DIR)
                _save_paths_as_presets(rule_paths, RULES_PRESET_DIR)

            # --- Presets selected from the library (reused, no need to re-upload) ---
            ref_presets = request.form.getlist('referencePresets')
            rule_presets = request.form.getlist('rulePresets')
            ref_paths += _resolve_preset_paths('reference', ref_presets, ALLOWED_EXTENSIONS)
            rule_paths += _resolve_preset_paths('rule', rule_presets, ALLOWED_RULE_EXTENSIONS)
            if ref_presets or rule_presets:
                print(f"[API] Dùng preset: {len(ref_presets)} sở cứ, {len(rule_presets)} quy định")

            # Create a dedicated analyzer for this session.
            # Empty reference sources → analyzer falls back to the default source (NĐ 86).
            # Empty rules → analyzer falls back to quy_dinh_chung.md.
            print("[API] Đang dựng index từ sở cứ...")
            analyzer = make_analyzer()

            if not analyzer.initialize_rag_system(
                ref_paths, rule_paths, history_paths, history_names
            ):
                return jsonify({
                    'error': 'Failed to initialize RAG system. Check logs for details.'
                }), 500
            
            # Run analysis
            print("[API] Running document analysis...")
            errors = analyzer.analyze_document(main_path)

            # Convert the main document to HTML for the preview (non-blocking on error)
            try:
                main_html = docx_to_html(main_path)
            except Exception as e:  # noqa: BLE001
                print(f"[API] Không tạo được HTML preview: {e}")
                main_html = ""

            # Store the session, including the analyzer for later cleanup
            _sessions[session_id] = {
                'main_path': main_path,
                'main_html': main_html,
                'ref_paths': ref_paths,
                'rule_paths': rule_paths,
                'history_paths': history_paths,
                'upload_dir': upload_dir,
                'analyzer': analyzer,
                'errors': errors,
                'error_count': len(errors)
            }
            
            print(f"[API] Analysis complete: {len(errors)} errors found")
            
            return jsonify({
                'success': True,
                'sessionId': session_id,
                'message': 'Document analyzed successfully',
                'results': {
                    'errorCount': len(errors),
                    'errors': errors
                }
            }), 200
        
        except Exception as e:
            print(f"[API] Error during analysis: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({
                'error': f'Analysis failed: {str(e)}'
            }), 500
    
    except Exception as e:
        print(f"[API] Upload error: {e}")
        return jsonify({'error': str(e)}), 500


@api_bp.route('/session/<session_id>', methods=['GET'])
def get_session_results(session_id):
    """
    Get analysis results for a session
    """
    if session_id not in _sessions:
        return jsonify({'error': 'Session not found'}), 404
    
    session = _sessions[session_id]
    return jsonify({
        'sessionId': session_id,
        'results': {
            'errorCount': session['error_count'],
            'errors': session['errors']
        }
    }), 200


@api_bp.route('/session/<session_id>/apply-suggestions', methods=['POST'])
def apply_suggestions(session_id):
    """
    Apply the accepted suggestions to the document and export the corrected DOCX file.

    Expected JSON:
    {
        "updates": [
            {
                "errorId": "error_c0_1",
                "action": "accept" | "reject" | "custom",
                "fixedValue": "..."   // optional: value entered manually by the user
            }
        ]
    }

    Response:
    {
        "success": true,
        "appliedCount": 3,
        "downloadUrl": "/api/session/<id>/download"
    }
    """
    try:
        if session_id not in _sessions:
            return jsonify({'error': 'Session not found'}), 404

        data = request.get_json()
        if not data or 'updates' not in data:
            return jsonify({'error': 'Trường "updates" là bắt buộc'}), 400

        session = _sessions[session_id]
        errors_by_id = {e['id']: e for e in session.get('errors', [])}
        updates = data['updates']

        # Build the list of text replacements for the accepted errors
        text_corrections = []
        for update in updates:
            if update.get('action') == 'reject':
                continue

            error = errors_by_id.get(update.get('errorId'))
            if not error:
                continue

            original_text = error.get('original_text', '').strip()
            if not original_text:
                continue

            # Priority: user's manually entered value > AI suggestion
            new_text = (
                update.get('fixedValue')
                or update.get('customSuggestion')
                or error.get('suggestion', '')
            ).strip()

            if new_text:
                text_corrections.append({
                    'original_text': original_text,
                    'new_text': new_text,
                })

        if not text_corrections:
            return jsonify({
                'success': True,
                'appliedCount': 0,
                'message': 'Không có sửa lỗi nào được chấp nhận.',
            }), 200

        # Apply to the DOCX file
        main_path = session['main_path']
        base_name = os.path.splitext(os.path.basename(main_path))[0]
        corrected_path = os.path.join(session['upload_dir'], f'{base_name}_corrected.docx')

        processor = DocumentProcessor()
        applied = processor.apply_text_corrections(main_path, corrected_path, text_corrections)

        session['corrected_path'] = corrected_path
        print(f"[API] Đã áp dụng {applied} sửa lỗi → {corrected_path}")

        return jsonify({
            'success': True,
            'appliedCount': applied,
            'downloadUrl': f'/api/session/{session_id}/download',
        }), 200

    except Exception as e:
        print(f"[API] Lỗi apply suggestions: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@api_bp.route('/session/<session_id>/download', methods=['GET'])
def download_corrected(session_id):
    """Download the corrected DOCX file."""
    if session_id not in _sessions:
        return jsonify({'error': 'Session not found'}), 404

    session = _sessions[session_id]
    corrected_path = session.get('corrected_path')

    if not corrected_path or not os.path.exists(corrected_path):
        return jsonify({
            'error': 'Chưa có file đã sửa. Hãy gọi apply-suggestions trước.'
        }), 404

    download_name = os.path.basename(corrected_path)
    return send_file(corrected_path, as_attachment=True, download_name=download_name)


@api_bp.route('/session/<session_id>/chat', methods=['POST'])
def chat(session_id):
    """
    Q&A with the chatbot inside DocumentPreview.

    The chatbot looks up information from PREVIOUSLY approved YCKT and can compare it
    with the DOCUMENT UNDER REVIEW (both sources are available; the LLM picks based on
    the question).

    Expected JSON:
    {
        "question": "So sánh van xả áp tài liệu này với YCKT trước đây",
        "history": [{"role": "user"|"assistant", "content": "..."}]   // optional
    }

    Response:
    {
        "success": true,
        "answer": "...",
        "citations": [
            {"source", "doc_name", "section", "param_name", "param_value", "score"}
        ]
    }
    """
    if session_id not in _sessions:
        return jsonify({'error': 'Session not found'}), 404

    data = request.get_json(silent=True) or {}
    question = (data.get('question') or '').strip()
    if not question:
        return jsonify({'error': 'Trường "question" là bắt buộc'}), 400

    history = data.get('history') or []

    analyzer = _sessions[session_id].get('analyzer')
    if analyzer is None:
        return jsonify({'error': 'Session chưa sẵn sàng (thiếu analyzer)'}), 409

    try:
        # Both sources (previous YCKT + the document under review) are available for the
        # chatbot to look up or cross-check depending on the question.
        result = analyzer.answer_question(question, history=history)
        return jsonify({'success': True, **result}), 200
    except Exception as e:
        print(f"[API] Lỗi chat: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@api_bp.route('/session/<session_id>/cleanup', methods=['DELETE'])
def cleanup_session(session_id):
    """
    Clean up session resources
    """
    if session_id in _sessions:
        session = _sessions.pop(session_id)

        # Close this session's Qdrant client
        analyzer = session.get('analyzer')
        if analyzer:
            analyzer.cleanup()

        # Remove the upload directory
        import shutil
        if os.path.exists(session['upload_dir']):
            try:
                shutil.rmtree(session['upload_dir'])
                print(f"[API] Đã xoá session {session_id}")
            except Exception:
                pass
    
    return jsonify({'success': True, 'message': 'Session cleaned up'}), 200


@api_bp.route('/rules/default', methods=['GET'])
def get_default_rules():
    """
    Return the default rules content (quy_dinh_chung.md) for display/reference in the
    UI. The user can preview it before deciding whether to upload an override.
    """
    try:
        from rag.knowledge_base.rules import load_rules
        return jsonify({
            'success': True,
            'rules': load_rules(),
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@api_bp.route('/session/<session_id>/document', methods=['GET'])
def get_document_html(session_id):
    """
    Return the main document's HTML + the error list, so the frontend can render the
    preview and highlight the incorrect passages (matched by original_text).
    """
    if session_id not in _sessions:
        return jsonify({'error': 'Session not found'}), 404
    session = _sessions[session_id]
    return jsonify({
        'success': True,
        'html': session.get('main_html', ''),
        'errors': session.get('errors', []),
    }), 200


@api_bp.route('/presets', methods=['GET'])
def list_presets():
    """List the saved reference-source & rule libraries available for reuse."""
    return jsonify({
        'success': True,
        'references': _list_presets('reference', ALLOWED_EXTENSIONS),
        'rules': _list_presets('rule', ALLOWED_RULE_EXTENSIONS),
    }), 200


@api_bp.route('/presets/<kind>/<path:name>', methods=['DELETE'])
def delete_preset(kind, name):
    """Delete a preset from the library (default presets cannot be deleted)."""
    if kind not in ('reference', 'rule'):
        return jsonify({'error': 'kind phải là reference hoặc rule'}), 400

    safe_name = os.path.basename(name)
    if safe_name in _PROTECTED_PRESETS:
        return jsonify({'error': f'Không thể xoá preset mặc định "{safe_name}"'}), 403

    target = _preset_dir(kind) / safe_name
    if not target.exists() or not target.is_file():
        return jsonify({'error': 'Preset không tồn tại'}), 404
    try:
        target.unlink()
        return jsonify({'success': True, 'message': f'Đã xoá {safe_name}'}), 200
    except Exception as e:  # noqa: BLE001
        return jsonify({'error': str(e)}), 500


@api_bp.route('/session/<session_id>/report.xlsx', methods=['GET'])
def export_report_excel(session_id):
    """Export the error-list report to an Excel file (.xlsx)."""
    if session_id not in _sessions:
        return jsonify({'error': 'Session not found'}), 404

    session = _sessions[session_id]
    errors = session.get('errors', [])

    try:
        xlsx_bytes = _build_excel_report(errors, os.path.basename(session['main_path']))
    except Exception as e:  # noqa: BLE001
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Không tạo được báo cáo Excel: {e}'}), 500

    return send_file(
        io.BytesIO(xlsx_bytes),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'bao_cao_tham_dinh_{session_id[:8]}.xlsx',
    )


def _build_excel_report(errors: list, doc_name: str) -> bytes:
    """
    Build the error-report workbook → bytes.

    Errors are GROUPED by "section" (heading/position within the audited document); each
    group has a header row so the reader can easily tell which part of the document an
    error belongs to. Each sub-error type is one row.
    """
    from collections import OrderedDict
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Báo cáo thẩm định"

    headers = [
        'STT', 'Nội dung gốc', 'Loại lỗi', 'Giải thích',
        'Đề xuất sửa', 'Vị trí sở cứ', 'Trích dẫn sở cứ',
    ]
    widths = [6, 40, 22, 50, 35, 24, 40]
    ncol = len(headers)

    # Document title
    ws.append([f'BÁO CÁO THẨM ĐỊNH: {doc_name}'])
    ws.append([f'Tổng số mục lỗi: {len(errors)}'])
    ws.append([])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)

    header_row_idx = ws.max_row + 1
    ws.append(headers)

    header_fill = PatternFill('solid', fgColor='4472C4')
    header_font = Font(bold=True, color='FFFFFF')
    section_fill = PatternFill('solid', fgColor='D9E1F2')
    section_font = Font(bold=True, color='1F3864', size=12)
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=header_row_idx, column=col_idx).column_letter].width = w
        c = ws.cell(row=header_row_idx, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border

    # Group errors by "section" (preserving their order of appearance in the document)
    groups = OrderedDict()
    for err in errors:
        sec = (err.get('section') or '').strip() or '(Không xác định mục)'
        groups.setdefault(sec, []).append(err)

    stt = 0
    for section, errs in groups.items():
        # Section header row (merged across the full width)
        sec_row = ws.max_row + 1
        ws.cell(row=sec_row, column=1, value=f'📂 Mục: {section}')
        ws.merge_cells(start_row=sec_row, start_column=1, end_row=sec_row, end_column=ncol)
        for col_idx in range(1, ncol + 1):
            c = ws.cell(row=sec_row, column=col_idx)
            c.fill = section_fill
            c.font = section_font
            c.border = border
        ws.cell(row=sec_row, column=1).alignment = Alignment(
            horizontal='left', vertical='center'
        )

        for err in errs:
            for sub in (err.get('danh_sach_cac_loi') or [{}]):
                stt += 1
                ws.append([
                    stt,
                    err.get('original_text', ''),
                    sub.get('error_type', ''),
                    sub.get('reasoning', ''),
                    err.get('suggestion', ''),
                    err.get('reference_location', ''),
                    err.get('reference_quote', ''),
                ])
                r = ws.max_row
                for col_idx in range(1, ncol + 1):
                    cell = ws.cell(row=r, column=col_idx)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    cell.border = border

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@api_bp.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({
        'status': 'ok',
        'sessions': len(_sessions)
    }), 200
